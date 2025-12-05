"""
Excel export utilities.

Functions for generating Excel files from data.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime
from fastapi.responses import Response


def generate_finance_xlsx(entries: List[Dict[str, Any]], target_date: str) -> bytes:
    """
    Generate Excel file for finance entries of a specific date.
    
    Args:
        entries: List of finance entry dictionaries
        target_date: Target date in YYYY-MM-DD format
    
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fluxo de Caixa"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Relatório de Fluxo de Caixa - {target_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = [
        "Data/Hora",
        "Tipo",
        "Valor",
        "Método de Pagamento",
        "Categoria",
        "Descrição"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 4
    total_income = 0.0
    total_expense = 0.0
    
    for entry in entries:
        # Extract date and time
        date_time = entry.get('date_time', '')
        if 'T' in date_time:
            date_part, time_part = date_time.split('T')
            formatted_datetime = f"{date_part} {time_part[:5]}"
        else:
            formatted_datetime = date_time
        
        entry_type = entry.get('type', '').upper()
        amount = float(entry.get('amount', 0))
        payment_method = entry.get('payment_method', '').upper()
        category = entry.get('category', '') or '-'
        description = entry.get('description', '') or '-'
        
        # Calculate totals
        if entry_type.lower() == 'income':
            total_income += amount
        else:
            total_expense += amount
        
        # Write row data
        ws.cell(row=row_num, column=1, value=formatted_datetime).border = border
        ws.cell(row=row_num, column=2, value=entry_type).border = border
        ws.cell(row=row_num, column=3, value=amount).border = border
        ws.cell(row=row_num, column=4, value=payment_method).border = border
        ws.cell(row=row_num, column=5, value=category).border = border
        ws.cell(row=row_num, column=6, value=description).border = border
        
        # Format amount column
        amount_cell = ws.cell(row=row_num, column=3)
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = Alignment(horizontal="right")
        
        row_num += 1
    
    # Summary section
    summary_row = row_num + 2
    ws.cell(row=summary_row, column=2, value="TOTAL RECEITAS:").font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=total_income).font = Font(bold=True)
    ws.cell(row=summary_row, column=3).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="right")
    
    summary_row += 1
    ws.cell(row=summary_row, column=2, value="TOTAL DESPESAS:").font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=total_expense).font = Font(bold=True)
    ws.cell(row=summary_row, column=3).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="right")
    
    summary_row += 1
    balance = total_income - total_expense
    ws.cell(row=summary_row, column=2, value="SALDO:").font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=3, value=balance).font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=3).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="right")
    
    # Color balance cell
    balance_cell = ws.cell(row=summary_row, column=3)
    if balance >= 0:
        balance_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        balance_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths
    column_widths = [20, 12, 15, 20, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Set row heights
    ws.row_dimensions[3].height = 20
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_classes_attendance_xlsx(
    classes: List[Dict[str, Any]], 
    attendance_records: List[Dict[str, Any]], 
    from_date: str, 
    to_date: str
) -> bytes:
    """
    Generate Excel file for classes and attendance records filtered by date range.
    
    Args:
        classes: List of class dictionaries
        attendance_records: List of attendance record dictionaries
        from_date: Start date in YYYY-MM-DD format
        to_date: End date in YYYY-MM-DD format
    
    Returns:
        Excel file as bytes
    """
    from app import crud
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all students for lookup
    all_students = crud.get_all_students()
    students_dict = {s['id']: s for s in all_students}
    
    # Create a sheet for each class
    for class_item in classes:
        class_id = class_item['id']
        class_name = class_item['name']
        class_date = class_item.get('date', '')
        class_time = class_item.get('time', '')
        
        # Get enrolled students for this class
        enrolled_students = crud.get_enrolled_students_by_class(class_id)
        enrolled_student_ids = {s['id'] for s in enrolled_students}
        
        # Create sheet for this class
        ws = wb.create_sheet(title=class_name[:31])  # Excel sheet name limit is 31 chars
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"{class_name} - {class_date} {class_time}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Class info
        ws.merge_cells('A2:E2')
        info_cell = ws['A2']
        info_cell.value = f"Período: {from_date} a {to_date}"
        info_cell.font = Font(size=10, italic=True)
        info_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["Aluno", "Data/Hora", "Status", "Observações"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Filter attendance for this class
        class_attendance = [
            att for att in attendance_records
            if att.get('class_id') == class_id
        ]
        
        # Create a dictionary of attendance by student_id for quick lookup
        attendance_by_student = {}
        for att_record in class_attendance:
            student_id = att_record.get('student_id')
            if student_id not in attendance_by_student:
                attendance_by_student[student_id] = []
            attendance_by_student[student_id].append(att_record)
        
        # Data rows - show all enrolled students
        row_num = 5
        present_count = 0
        absent_count = 0
        
        # Sort enrolled students by name
        enrolled_students_sorted = sorted(enrolled_students, key=lambda x: x.get('name', ''))
        
        for student in enrolled_students_sorted:
            student_id = student['id']
            student_name = student.get('name', 'Aluno não encontrado')
            
            # Get attendance records for this student in this class
            student_attendance = attendance_by_student.get(student_id, [])
            
            if student_attendance:
                # Show each attendance record
                for att_record in sorted(student_attendance, key=lambda x: x.get('date_time', '')):
                    # Format date_time
                    date_time = att_record.get('date_time', '')
                    if 'T' in date_time:
                        date_part, time_part = date_time.split('T')
                        formatted_datetime = f"{date_part} {time_part[:5]}"
                    else:
                        formatted_datetime = date_time
                    
                    status = att_record.get('status', '').upper()
                    if status == 'PRESENT':
                        present_count += 1
                        status_text = 'Presente'
                    else:
                        absent_count += 1
                        status_text = 'Falta'
                    
                    # Write row data
                    ws.cell(row=row_num, column=1, value=student_name).border = border
                    ws.cell(row=row_num, column=2, value=formatted_datetime).border = border
                    ws.cell(row=row_num, column=3, value=status_text).border = border
                    ws.cell(row=row_num, column=4, value='').border = border
                    
                    row_num += 1
            else:
                # Student enrolled but no attendance recorded
                ws.cell(row=row_num, column=1, value=student_name).border = border
                ws.cell(row=row_num, column=2, value='-').border = border
                ws.cell(row=row_num, column=3, value='Sem registro').border = border
                ws.cell(row=row_num, column=4, value='').border = border
                
                row_num += 1
        
        # Summary section
        summary_row = row_num + 2
        ws.cell(row=summary_row, column=1, value="RESUMO:").font = Font(bold=True, size=12)
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Total de Presenças:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=present_count).font = Font(bold=True)
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Total de Faltas:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=absent_count).font = Font(bold=True)
        
        summary_row += 1
        total = present_count + absent_count
        attendance_rate = (present_count / total * 100) if total > 0 else 0
        ws.cell(row=summary_row, column=1, value="Taxa de Presença:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=f"{attendance_rate:.1f}%").font = Font(bold=True)
        
        # Auto-adjust column widths
        column_widths = [40, 20, 15, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Set row heights
        ws.row_dimensions[4].height = 20
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title="Resumo", index=0)
    
    # Title
    summary_ws.merge_cells('A1:F1')
    title_cell = summary_ws['A1']
    title_cell.value = f"Relatório de Aulas e Presença - {from_date} a {to_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 25
    
    # Headers
    summary_headers = ["Aula", "Data", "Hora", "Total Alunos", "Presentes", "Faltas", "Taxa Presença"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Summary data
    row_num = 4
    for class_item in classes:
        class_id = class_item['id']
        class_name = class_item['name']
        class_date = class_item.get('date', '')
        class_time = class_item.get('time', '')
        
        # Get enrolled students count
        enrolled_students = crud.get_enrolled_students_by_class(class_id)
        total_enrolled = len(enrolled_students)
        
        # Count attendance for this class
        class_attendance = [
            att for att in attendance_records
            if att.get('class_id') == class_id
        ]
        
        present_count = sum(1 for att in class_attendance if att.get('status', '').lower() == 'present')
        absent_count = sum(1 for att in class_attendance if att.get('status', '').lower() == 'absent')
        total_attendance = present_count + absent_count
        attendance_rate = (present_count / total_attendance * 100) if total_attendance > 0 else 0
        
        summary_ws.cell(row=row_num, column=1, value=class_name).border = border
        summary_ws.cell(row=row_num, column=2, value=class_date).border = border
        summary_ws.cell(row=row_num, column=3, value=class_time).border = border
        summary_ws.cell(row=row_num, column=4, value=total_enrolled).border = border
        summary_ws.cell(row=row_num, column=5, value=present_count).border = border
        summary_ws.cell(row=row_num, column=6, value=absent_count).border = border
        summary_ws.cell(row=row_num, column=7, value=f"{attendance_rate:.1f}%").border = border
        
        row_num += 1
    
    # Auto-adjust column widths for summary
    summary_widths = [30, 12, 10, 12, 12, 12, 15]
    for col_num, width in enumerate(summary_widths, 1):
        summary_ws.column_dimensions[get_column_letter(col_num)].width = width
    
    summary_ws.row_dimensions[3].height = 20
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()